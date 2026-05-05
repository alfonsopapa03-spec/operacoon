import streamlit as st
import psycopg2
from psycopg2 import pool as pg_pool
import pandas as pd
from datetime import datetime, timedelta, date
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pytz

# ==================== CONFIGURACIÓN ====================
st.set_page_config(
    page_title="Inspecciones Preoperacionales",
    layout="wide",
    page_icon="🔧",
    initial_sidebar_state="collapsed"
)

# ==================== CREDENCIALES ====================
SUPABASE_DB_URL = "postgresql://postgres.ogfenizdijcboekqhuhd:Conejito200$@aws-1-us-west-2.pooler.supabase.com:6543/postgres"

# ==================== CATÁLOGO DE MÁQUINAS ====================
MAQUINAS = [
    "Bloquera",
    "Caldera",
    "Dobladora",
    "Enderezadora",
    "Enmalladora",
    "Molino",
    "Montacarga",
    "Paneladora",
    "Pantógrafo",
    "Preexpansora",
    "Soldador Manual",
]

# ==================== DÍAS DE LA SEMANA ====================
DIAS_SEMANA = ["Lun", "Mar", "Mier", "Juev", "Vier", "Sáb", "Dom"]

# ==================== ÍTEMS DE INSPECCIÓN ====================
ITEMS_ANTES_USO = [
    "¿Tiene permiso el trabajador para utilizar la máquina?",
    "¿Ha sido capacitado el trabajador para utilizar la máquina?",
    "¿Se ha verificado que la presión del aire se encuentre en 125 PSI?",
    "¿Se ha inspeccionado que los desviadores contengan material adecuadamente?",
    "¿Se ha inspeccionado que los electro-válvulas funcionen adecuadamente?",
    "¿Se ha comprobado que los ganchos de ajuste funcionen correctamente?",
    "¿El 'carrusel' de fricción del material funciona satisfactoriamente?",
    "¿Se ha verificado que el tope se encuentre en óptimas condiciones?",
    "¿Se ha inspeccionado que los botones y controles funcionen oportunamente?",
    "¿Las paradas de emergencia (6 en total) funcionan correctamente?",
    "¿Se ha verificado el estado de los cabezales (inferior/superior)?",
    "¿El nivel de aceite de lubricación se encuentra en nivel adecuado?",
    "¿El manómetro de aceite de sobrecarga funciona correctamente?",
    "¿Se ha ajustado la altura del panel a la medida correspondiente?",
]

ITEMS_EPP = [
    "¿Se ha inspeccionado el lugar de trabajo? (material combustible, riesgo de incendios, instalaciones, otros trabajadores, etc.)",
    "¿La iluminación del área de trabajo es adecuada para operación de la máquina sin riesgos?",
    "¿Cuenta con los elementos de protección personal? (protector de ojos, oídos, guantes y cabezado)",
    "¿El trabajador está vestido apropiadamente? (Camisa manga larga, pantalón de dotación y calzado de seguridad)",
    "¿Se evidencia el NO uso de joyas, relojes y ropa holgada?",
    "¿Se tiene el cabello recogido si lo tiene largo?",
]

ITEMS_ELECTRICA = [
    "¿Se ha verificado que el cable de alimentación está en buen estado?",
    "¿Se ha revisado que el enchufe se encuentre en buenas condiciones?",
    "¿El interruptor de encendido funciona correctamente?",
]

TODOS_ITEMS = ITEMS_ANTES_USO + ITEMS_EPP + ITEMS_ELECTRICA
OPCIONES_INSPECCION = ["C", "NC", "N/A"]
ESTADOS_INSPECCION  = ["✅ Aprobada", "⚠️ Con Observaciones", "❌ Rechazada"]

# ==================== CSS ====================
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@400;600;700&family=Barlow:wght@300;400;500&display=swap');
    html, body, [class*="css"] { font-family: 'Barlow', sans-serif; }
    .main-header {
        background: linear-gradient(135deg, #0f2027, #203a43, #2c5364);
        padding: 1.5rem 2rem; border-radius: 12px; margin-bottom: 1.5rem;
    }
    .main-header h1 {
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 2rem; font-weight: 700; color: white; margin: 0; letter-spacing: 1px;
    }
    .main-header p { color: #a0c4d8; margin: 0; font-size: 0.9rem; }
    .seccion-titulo {
        background: #203a43; color: white; padding: 0.4rem 1rem;
        border-radius: 6px; font-weight: 700; font-size: 0.95rem;
        margin: 1rem 0 0.5rem 0; letter-spacing: 0.5px;
    }
    .item-label { font-size: 0.85rem; color: #333; padding: 0.3rem 0; }
    .badge-c   { background: #2ecc71; color: white; border-radius: 4px; padding: 2px 8px; font-weight: 700; font-size: 0.8rem; }
    .badge-nc  { background: #e74c3c; color: white; border-radius: 4px; padding: 2px 8px; font-weight: 700; font-size: 0.8rem; }
    .badge-na  { background: #95a5a6; color: white; border-radius: 4px; padding: 2px 8px; font-weight: 700; font-size: 0.8rem; }
    div[data-testid="stTabs"] button {
        font-family: 'Barlow Condensed', sans-serif;
        font-weight: 600; font-size: 1rem; letter-spacing: 0.5px;
    }
    .kpi-box {
        background: white; border-radius: 10px; padding: 1rem 1.2rem;
        border-left: 5px solid #2c5364; box-shadow: 0 2px 8px rgba(0,0,0,0.07);
        margin-bottom: 0.5rem;
    }
    .campo-obligatorio { color: #e74c3c; font-weight: bold; }
    .dia-header {
        background: #2c5364; color: white; text-align: center;
        padding: 4px 0; border-radius: 4px; font-weight: 700;
        font-size: 0.8rem; margin-bottom: 2px;
    }
    .tabla-semanal th {
        background: #203a43 !important; color: white !important;
        text-align: center; padding: 6px;
    }
    .tabla-semanal td { padding: 4px 6px; vertical-align: middle; }
</style>
""", unsafe_allow_html=True)


# ==================== BASE DE DATOS ====================
@st.cache_resource
def get_pool():
    try:
        connection_pool = pg_pool.SimpleConnectionPool(
            minconn=1,
            maxconn=5,
            dsn=SUPABASE_DB_URL,
            sslmode="require",
            connect_timeout=15,
            options="-c statement_timeout=30000"
        )
        return connection_pool
    except Exception as e:
        st.error(f"❌ No se pudo conectar a la base de datos: {e}")
        st.stop()


class DB:
    def __init__(self):
        self.pool = get_pool()
        self.init()

    def conn(self):
        try:
            c = self.pool.getconn()
            c.cursor().execute("SELECT 1")
            return c
        except Exception:
            try:
                return psycopg2.connect(
                    dsn=SUPABASE_DB_URL,
                    sslmode="require",
                    connect_timeout=15,
                )
            except Exception as e:
                st.error(f"❌ Error de conexión: {e}")
                st.stop()

    def release(self, c):
        try:
            if c and not c.closed:
                self.pool.putconn(c)
        except Exception:
            pass

    def init(self):
        c = None
        try:
            c = self.conn()
            cur = c.cursor()

            # ── Detectar columnas existentes ──────────────────────────────────
            cur.execute("""
                SELECT column_name FROM information_schema.columns
                WHERE table_name = 'inspecciones_preop'
            """)
            cols_existentes = {row[0] for row in cur.fetchall()}

            if not cols_existentes:
                # Tabla nueva: crear sin columna 'fecha'
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS inspecciones_preop (
                        id SERIAL PRIMARY KEY,
                        fecha_registro TIMESTAMP DEFAULT (now() AT TIME ZONE 'America/Bogota'),
                        fecha_inicio DATE NOT NULL,
                        fecha_fin DATE NOT NULL,
                        maquina TEXT NOT NULL,
                        modelo TEXT,
                        marca TEXT,
                        placa TEXT,
                        trabajador TEXT NOT NULL,
                        revisado_por TEXT NOT NULL,
                        cliente_proyecto TEXT NOT NULL,
                        responsable_mantenimiento TEXT NOT NULL,
                        estado TEXT DEFAULT 'Aprobada',
                        observaciones TEXT
                    )
                """)
            else:
                # Tabla existente: agregar columnas faltantes
                if "fecha_inicio" not in cols_existentes:
                    cur.execute("ALTER TABLE inspecciones_preop ADD COLUMN fecha_inicio DATE")
                    # Si existía 'fecha', copiar sus valores
                    if "fecha" in cols_existentes:
                        cur.execute("UPDATE inspecciones_preop SET fecha_inicio = fecha WHERE fecha_inicio IS NULL")
                    else:
                        cur.execute("UPDATE inspecciones_preop SET fecha_inicio = CURRENT_DATE WHERE fecha_inicio IS NULL")
                if "fecha_fin" not in cols_existentes:
                    cur.execute("ALTER TABLE inspecciones_preop ADD COLUMN fecha_fin DATE")
                    if "fecha" in cols_existentes:
                        cur.execute("UPDATE inspecciones_preop SET fecha_fin = fecha WHERE fecha_fin IS NULL")
                    else:
                        cur.execute("UPDATE inspecciones_preop SET fecha_fin = CURRENT_DATE WHERE fecha_fin IS NULL")

            # ── Tabla de ítems ────────────────────────────────────────────────
            cur.execute("""
                CREATE TABLE IF NOT EXISTS inspecciones_preop_items (
                    id SERIAL PRIMARY KEY,
                    inspeccion_id INTEGER REFERENCES inspecciones_preop(id) ON DELETE CASCADE,
                    seccion TEXT NOT NULL,
                    item_numero INTEGER NOT NULL,
                    descripcion TEXT NOT NULL,
                    dia TEXT NOT NULL DEFAULT 'General',
                    resultado TEXT DEFAULT 'C'
                )
            """)
            cur.execute("""
                SELECT column_name FROM information_schema.columns
                WHERE table_name = 'inspecciones_preop_items'
            """)
            item_cols = {row[0] for row in cur.fetchall()}
            if "dia" not in item_cols:
                cur.execute("ALTER TABLE inspecciones_preop_items ADD COLUMN dia TEXT NOT NULL DEFAULT 'General'")

            c.commit()
            cur.close()
        except Exception as e:
            st.error(f"Error inicializando DB: {e}")
        finally:
            self.release(c)

    def guardar_inspeccion(self, datos: dict, items: list) -> bool:
        c = None
        try:
            c = self.conn()
            cur = c.cursor()
            cur.execute("""
                INSERT INTO inspecciones_preop
                (fecha_inicio, fecha_fin, maquina, modelo, marca, placa, trabajador, revisado_por,
                 cliente_proyecto, responsable_mantenimiento, estado, observaciones)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                RETURNING id
            """, (
                datos["fecha_inicio"], datos["fecha_fin"],
                datos["maquina"], datos["modelo"], datos["marca"],
                datos["placa"], datos["trabajador"], datos["revisado_por"],
                datos["cliente_proyecto"], datos["responsable_mantenimiento"],
                datos["estado"], datos["observaciones"],
            ))
            inspeccion_id = cur.fetchone()[0]
            for item in items:
                cur.execute("""
                    INSERT INTO inspecciones_preop_items
                    (inspeccion_id, seccion, item_numero, descripcion, dia, resultado)
                    VALUES (%s,%s,%s,%s,%s,%s)
                """, (
                    inspeccion_id, item["seccion"], item["item_numero"],
                    item["descripcion"], item["dia"], item["resultado"]
                ))
            c.commit()
            cur.close()
            return True
        except Exception as e:
            st.error(f"Error guardando: {e}")
            return False
        finally:
            self.release(c)

    def actualizar_inspeccion(self, inspeccion_id: int, datos: dict, items: list) -> bool:
        c = None
        try:
            c = self.conn()
            cur = c.cursor()
            cur.execute("""
                UPDATE inspecciones_preop SET
                fecha_inicio=%s, fecha_fin=%s,
                maquina=%s, modelo=%s, marca=%s, placa=%s,
                trabajador=%s, revisado_por=%s, cliente_proyecto=%s,
                responsable_mantenimiento=%s, estado=%s, observaciones=%s
                WHERE id=%s
            """, (
                datos["fecha_inicio"], datos["fecha_fin"],
                datos["maquina"], datos["modelo"], datos["marca"],
                datos["placa"], datos["trabajador"], datos["revisado_por"],
                datos["cliente_proyecto"], datos["responsable_mantenimiento"],
                datos["estado"], datos["observaciones"], inspeccion_id
            ))
            cur.execute(
                "DELETE FROM inspecciones_preop_items WHERE inspeccion_id=%s",
                (inspeccion_id,)
            )
            for item in items:
                cur.execute("""
                    INSERT INTO inspecciones_preop_items
                    (inspeccion_id, seccion, item_numero, descripcion, dia, resultado)
                    VALUES (%s,%s,%s,%s,%s,%s)
                """, (
                    inspeccion_id, item["seccion"], item["item_numero"],
                    item["descripcion"], item["dia"], item["resultado"]
                ))
            c.commit()
            cur.close()
            return True
        except Exception as e:
            st.error(f"Error actualizando: {e}")
            return False
        finally:
            self.release(c)

    def eliminar_inspeccion(self, inspeccion_id: int) -> bool:
        c = None
        try:
            c = self.conn()
            cur = c.cursor()
            cur.execute("DELETE FROM inspecciones_preop WHERE id=%s", (inspeccion_id,))
            c.commit()
            cur.close()
            return True
        except Exception as e:
            st.error(f"Error eliminando: {e}")
            return False
        finally:
            self.release(c)

    def obtener_inspecciones(self, fecha_ini=None, fecha_fin=None,
                              maquina=None, estado=None, trabajador=None) -> pd.DataFrame:
        c = None
        try:
            c = self.conn()
            q = """
                SELECT id,
                       fecha_inicio,
                       fecha_fin,
                       maquina, modelo, marca, placa,
                       trabajador, revisado_por, cliente_proyecto,
                       responsable_mantenimiento, estado, observaciones,
                       fecha_registro
                FROM inspecciones_preop WHERE 1=1
            """
            params = []
            if fecha_ini:
                q += " AND fecha_inicio >= %s"; params.append(fecha_ini)
            if fecha_fin:
                q += " AND fecha_fin <= %s"; params.append(fecha_fin)
            if maquina and maquina != "Todas":
                q += " AND maquina = %s"; params.append(maquina)
            if estado and estado != "Todos":
                q += " AND estado ILIKE %s"; params.append(f"%{estado}%")
            if trabajador:
                q += " AND trabajador ILIKE %s"; params.append(f"%{trabajador}%")
            q += " ORDER BY fecha_inicio DESC, id DESC"
            return pd.read_sql(q, c, params=params)
        except Exception as e:
            st.error(f"Error consultando inspecciones: {e}")
            return pd.DataFrame()
        finally:
            self.release(c)

    def obtener_items_inspeccion(self, inspeccion_id: int) -> pd.DataFrame:
        c = None
        try:
            c = self.conn()
            return pd.read_sql("""
                SELECT seccion, item_numero, descripcion,
                       COALESCE(dia, 'General') AS dia, resultado
                FROM inspecciones_preop_items
                WHERE inspeccion_id = %s
                ORDER BY seccion, item_numero, dia
            """, c, params=[inspeccion_id])
        except Exception as e:
            st.error(f"Error obteniendo ítems: {e}")
            return pd.DataFrame()
        finally:
            self.release(c)

    def obtener_todos_los_items(self, ids: list) -> pd.DataFrame:
        if not ids:
            return pd.DataFrame()
        c = None
        try:
            c = self.conn()
            return pd.read_sql("""
                SELECT inspeccion_id, seccion, item_numero, descripcion,
                       COALESCE(dia, 'General') AS dia, resultado
                FROM inspecciones_preop_items
                WHERE inspeccion_id = ANY(%s)
                ORDER BY inspeccion_id, seccion, item_numero, dia
            """, c, params=[ids])
        except Exception as e:
            st.error(f"Error obteniendo todos los ítems: {e}")
            return pd.DataFrame()
        finally:
            self.release(c)

    def stats_dashboard(self, fecha_ini, fecha_fin) -> pd.DataFrame:
        c = None
        try:
            c = self.conn()
            return pd.read_sql("""
                SELECT i.id,
                       i.fecha_inicio AS fecha,
                       i.maquina, i.trabajador, i.estado,
                       COUNT(CASE WHEN it.resultado = 'NC' THEN 1 END) as num_nc,
                       COUNT(CASE WHEN it.resultado = 'C'  THEN 1 END) as num_c,
                       COUNT(it.id) as total_items
                FROM inspecciones_preop i
                LEFT JOIN inspecciones_preop_items it ON it.inspeccion_id = i.id
                WHERE i.fecha_inicio >= %s
                  AND i.fecha_inicio <= %s
                GROUP BY i.id, i.fecha_inicio, i.maquina, i.trabajador, i.estado
                ORDER BY i.fecha_inicio
            """, c, params=[fecha_ini, fecha_fin])
        except Exception as e:
            st.error(f"Error en stats: {e}")
            return pd.DataFrame()
        finally:
            self.release(c)

    def verificar_inspeccion_existente(self, fecha_inicio, maquina) -> bool:
        c = None
        try:
            c = self.conn()
            cur = c.cursor()
            cur.execute(
                "SELECT COUNT(*) FROM inspecciones_preop WHERE fecha_inicio=%s AND maquina=%s",
                (fecha_inicio, maquina)
            )
            count = cur.fetchone()[0]
            cur.close()
            return count > 0
        except Exception:
            return False
        finally:
            self.release(c)


# ==================== HELPERS ====================

def construir_items_semanal(prefix: str, dias_activos: list) -> list:
    """
    Construye la lista de ítems considerando cada día activo de la semana.
    Genera un ítem por (descripción, día).
    """
    items = []
    secciones = [
        ("ANTES DE SU USO",                  ITEMS_ANTES_USO, "au"),
        ("ELEMENTOS DE PROTECCIÓN PERSONAL",  ITEMS_EPP,       "epp"),
        ("SEGURIDAD ELÉCTRICA",               ITEMS_ELECTRICA, "elec"),
    ]
    for sec_nombre, sec_items, sec_key in secciones:
        for i, desc in enumerate(sec_items):
            for dia in dias_activos:
                key = f"{prefix}_{sec_key}_{i}_{dia}"
                resultado = st.session_state.get(key, "C")
                items.append({
                    "seccion":     sec_nombre,
                    "item_numero": i + 1,
                    "descripcion": desc,
                    "dia":         dia,
                    "resultado":   resultado,
                })
    return items


def badge_resultado(val):
    if val == "C":   return "🟢 C"
    if val == "NC":  return "🔴 NC"
    return "⚪ N/A"


def render_tabla_semanal(seccion_label: str, items_lista: list, prefix: str,
                          sec_key: str, dias_activos: list, valores_previos: dict = None):
    """
    Renderiza una tabla semanal: filas = ítems, columnas = días activos.
    Cada celda tiene un selectbox C / NC / N/A.
    """
    st.markdown(f"<div class='seccion-titulo'>📋 {seccion_label}</div>", unsafe_allow_html=True)
    st.caption("**C** = Cumple · **NC** = No Cumple · **N/A** = No Aplica")

    if not dias_activos:
        st.warning("Selecciona al menos un día activo para mostrar la tabla.")
        return

    # Encabezados
    num_dias = len(dias_activos)
    # col widths: descripción (ancha) + una por día
    col_ratios = [4] + [1] * num_dias
    header_cols = st.columns(col_ratios)
    header_cols[0].markdown("**Ítem**")
    for j, dia in enumerate(dias_activos):
        header_cols[j + 1].markdown(
            f"<div class='dia-header'>{dia}</div>", unsafe_allow_html=True
        )

    # Filas de ítems
    for i, desc in enumerate(items_lista):
        row_cols = st.columns(col_ratios)
        with row_cols[0]:
            st.markdown(
                f"<div class='item-label'>{i+1}. {desc}</div>",
                unsafe_allow_html=True
            )
        for j, dia in enumerate(dias_activos):
            key  = f"{prefix}_{sec_key}_{i}_{dia}"
            prev = "C"
            if valores_previos:
                prev = valores_previos.get(key, "C")
            with row_cols[j + 1]:
                st.selectbox(
                    "R", OPCIONES_INSPECCION,
                    index=OPCIONES_INSPECCION.index(prev) if prev in OPCIONES_INSPECCION else 0,
                    key=key,
                    label_visibility="collapsed"
                )


def validar_datos_control(trabajador, revisado_por, cliente_proyecto, resp_mantenimiento) -> list:
    errores = []
    if not trabajador or not trabajador.strip():
        errores.append("👷 **Trabajador** es obligatorio.")
    if not revisado_por or not revisado_por.strip():
        errores.append("👤 **Revisado por** es obligatorio.")
    if not cliente_proyecto or not cliente_proyecto.strip():
        errores.append("🏢 **Cliente / Proyecto** es obligatorio.")
    if not resp_mantenimiento or not resp_mantenimiento.strip():
        errores.append("🔧 **Responsable de Mantenimiento** es obligatorio.")
    return errores


# ==================== EXCEL ====================
def generar_excel(df_inspecciones: pd.DataFrame, db: "DB", titulo: str = "Inspecciones Preoperacionales") -> bytes:
    wb = Workbook()

    ft_titulo = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ft_header = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    ft_normal = Font(name="Calibri", size=9)
    ft_total  = Font(name="Calibri", bold=True, size=10)
    ft_nc     = Font(name="Calibri", size=9, color="C0392B", bold=True)
    ft_apro   = Font(name="Calibri", size=9, color="1E8449")

    fill_titulo  = PatternFill("solid", start_color="0F2027")
    fill_header  = PatternFill("solid", start_color="203A43")
    fill_alt     = PatternFill("solid", start_color="EBF5FB")
    fill_total   = PatternFill("solid", start_color="D5DBDB")
    fill_nc_row  = PatternFill("solid", start_color="FADBD8")
    fill_obs_row = PatternFill("solid", start_color="FDEBD0")

    borde  = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"),  bottom=Side(style="thin"))
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    izq    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    now_col = datetime.now(pytz.timezone("America/Bogota"))

    ids_list = df_inspecciones["id"].astype(int).tolist()
    df_all_items = db.obtener_todos_los_items(ids_list)
    items_por_id = {}
    if not df_all_items.empty:
        for insp_id, grp in df_all_items.groupby("inspeccion_id"):
            items_por_id[int(insp_id)] = grp

    # ── HOJA 1: RESUMEN ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Inspecciones"
    total_cols = 15

    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    ws["A1"] = f"🔧 {titulo}   |   Generado: {now_col.strftime('%d/%m/%Y %H:%M')} (COL)   |   Total: {len(df_inspecciones)} inspecciones"
    ws["A1"].font      = ft_titulo
    ws["A1"].fill      = fill_titulo
    ws["A1"].alignment = centro
    ws.row_dimensions[1].height = 30

    columnas = [
        ("id",                        "ID",            6),
        ("fecha_inicio",              "F. INICIO",     12),
        ("fecha_fin",                 "F. FIN",        12),
        ("maquina",                   "MÁQUINA",       20),
        ("modelo",                    "MODELO",        14),
        ("marca",                     "MARCA",         14),
        ("placa",                     "PLACA",         12),
        ("trabajador",                "TRABAJADOR",    24),
        ("revisado_por",              "REVISADO POR",  24),
        ("cliente_proyecto",          "CLIENTE/PROY.", 20),
        ("responsable_mantenimiento", "RESP. MANT.",   24),
        ("estado",                    "ESTADO",        18),
        ("num_nc",                    "# NC",           7),
        ("num_c",                     "# C",            7),
        ("observaciones",             "OBSERVACIONES", 32),
    ]

    for idx, (key, nombre, ancho) in enumerate(columnas, start=1):
        cell = ws.cell(row=2, column=idx, value=nombre)
        cell.font = ft_header; cell.fill = fill_header
        cell.alignment = centro; cell.border = borde
        ws.column_dimensions[get_column_letter(idx)].width = ancho
    ws.row_dimensions[2].height = 28

    for row_idx, (_, fila) in enumerate(df_inspecciones.iterrows(), start=3):
        insp_id  = int(fila["id"])
        df_it    = items_por_id.get(insp_id, pd.DataFrame())
        num_nc   = len(df_it[df_it["resultado"] == "NC"]) if not df_it.empty else 0
        num_c    = len(df_it[df_it["resultado"] == "C"])  if not df_it.empty else 0
        estado_val = str(fila.get("estado", ""))
        es_rech  = "Rechazada"      in estado_val
        es_obs   = "Observaciones"  in estado_val
        fill_f   = fill_nc_row if es_rech else (fill_obs_row if es_obs else (fill_alt if row_idx % 2 == 0 else None))

        fi_val = str(fila.get("fecha_inicio", ""))
        ff_val = str(fila.get("fecha_fin",    ""))

        valores = {
            "id": insp_id,
            "fecha_inicio": fi_val,
            "fecha_fin": ff_val,
            "maquina": fila.get("maquina",""), "modelo": fila.get("modelo",""),
            "marca": fila.get("marca",""), "placa": fila.get("placa",""),
            "trabajador": fila.get("trabajador",""), "revisado_por": fila.get("revisado_por",""),
            "cliente_proyecto": fila.get("cliente_proyecto",""),
            "responsable_mantenimiento": fila.get("responsable_mantenimiento",""),
            "estado": estado_val, "num_nc": num_nc, "num_c": num_c,
            "observaciones": fila.get("observaciones",""),
        }

        for col_idx, (key, _, _) in enumerate(columnas, start=1):
            val  = valores.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val != "" else "")
            cell.border = borde
            cell.alignment = centro if key in ("id","fecha_inicio","fecha_fin","estado","num_nc","num_c","placa") else izq
            if key == "num_nc" and num_nc > 0:
                cell.font = ft_nc
            elif key == "estado" and "Aprobada" in str(val):
                cell.font = ft_apro
            else:
                cell.font = ft_normal
            if fill_f:
                cell.fill = fill_f
        ws.row_dimensions[row_idx].height = 18

    aprobadas = len(df_inspecciones[df_inspecciones["estado"].str.contains("Aprobada",      na=False)])
    obs_count = len(df_inspecciones[df_inspecciones["estado"].str.contains("Observaciones", na=False)])
    rech      = len(df_inspecciones[df_inspecciones["estado"].str.contains("Rechazada",     na=False)])
    total_row = len(df_inspecciones) + 3
    try:
        ws.merge_cells(f"A{total_row}:{get_column_letter(total_cols)}{total_row}")
    except Exception:
        pass
    ct = ws.cell(row=total_row, column=1,
                 value=f"TOTAL: {len(df_inspecciones)}   |   ✅ Aprobadas: {aprobadas}   ⚠️ Con Obs: {obs_count}   ❌ Rechazadas: {rech}")
    ct.font = ft_total; ct.fill = fill_total; ct.alignment = centro
    ws.freeze_panes = "A3"

    # ── HOJA 2: DETALLE ÍTEMS (con columna DÍA) ──────────────────────────────
    ws2 = wb.create_sheet("Detalle Ítems")
    ws2.merge_cells("A1:H1")
    ws2["A1"] = "Detalle de Ítems por Inspección (Semanal)"
    ws2["A1"].font = ft_titulo; ws2["A1"].fill = fill_titulo; ws2["A1"].alignment = centro
    ws2.row_dimensions[1].height = 26

    hdrs2   = ["ID INSP.", "F. INICIO", "F. FIN", "MÁQUINA", "SECCIÓN", "N°", "DESCRIPCIÓN DEL ÍTEM", "DÍA", "RESULTADO"]
    anchos2 = [8, 12, 12, 20, 30, 5, 65, 8, 10]
    for ci, (h, w) in enumerate(zip(hdrs2, anchos2), start=1):
        c = ws2.cell(2, ci, h)
        c.font = ft_header; c.fill = fill_header; c.alignment = centro; c.border = borde
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[2].height = 24

    fila2 = 3
    for _, insp in df_inspecciones.iterrows():
        insp_id = int(insp["id"])
        df_it   = items_por_id.get(insp_id, pd.DataFrame())
        if df_it.empty:
            continue
        fi_v = str(insp.get("fecha_inicio", ""))
        ff_v = str(insp.get("fecha_fin",    ""))
        for _, item in df_it.iterrows():
            res = str(item.get("resultado", "C"))
            fill_item = PatternFill("solid", start_color="FADBD8") if res == "NC" else (
                        PatternFill("solid", start_color="EBF5FB") if fila2 % 2 == 0 else None)
            vals = [insp_id, fi_v, ff_v, str(insp.get("maquina","")),
                    str(item.get("seccion","")), int(item.get("item_numero",0)),
                    str(item.get("descripcion","")), str(item.get("dia","General")), res]
            for ci, v in enumerate(vals, start=1):
                c = ws2.cell(fila2, ci, v)
                c.font = ft_nc if res == "NC" else ft_normal
                c.border = borde
                c.alignment = izq if ci == 7 else centro
                if fill_item:
                    c.fill = fill_item
            ws2.row_dimensions[fila2].height = 18
            fila2 += 1
    ws2.freeze_panes = "A3"

    # ── HOJA 3: POR MÁQUINA ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("Por Máquina")
    ws3.merge_cells("A1:H1")
    ws3["A1"] = "Resumen de Inspecciones por Máquina"
    ws3["A1"].font = ft_titulo; ws3["A1"].fill = fill_titulo; ws3["A1"].alignment = centro
    ws3.row_dimensions[1].height = 26

    hdrs3   = ["MÁQUINA","TOTAL INSP.","✅ APROBADAS","⚠️ CON OBS.","❌ RECHAZADAS","% APROBACIÓN","ÚLTIMO INSPECTOR","ÚLTIMA SEMANA"]
    anchos3 = [22, 12, 14, 14, 14, 14, 28, 14]
    for ci, (h, w) in enumerate(zip(hdrs3, anchos3), start=1):
        c = ws3.cell(2, ci, h)
        c.font = ft_header; c.fill = fill_header; c.alignment = centro; c.border = borde
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.row_dimensions[2].height = 24

    if not df_inspecciones.empty and "maquina" in df_inspecciones.columns:
        fecha_col = "fecha_inicio"
        resumen_maq = df_inspecciones.groupby("maquina", as_index=False).agg(
            total       =("maquina",  "count"),
            aprobadas   =("estado",   lambda x: x.str.contains("Aprobada",      na=False).sum()),
            con_obs     =("estado",   lambda x: x.str.contains("Observaciones", na=False).sum()),
            rechazadas  =("estado",   lambda x: x.str.contains("Rechazada",     na=False).sum()),
            ultima_fecha=(fecha_col,  "max"),
        ).sort_values("total", ascending=False)

        ultimo_insp_map = (
            df_inspecciones.sort_values(fecha_col)
            .groupby("maquina")["trabajador"]
            .last()
            .reset_index()
            .rename(columns={"trabajador": "ultimo_insp"})
        )
        resumen_maq = resumen_maq.merge(ultimo_insp_map, on="maquina", how="left")
        resumen_maq["ultima_fecha"] = resumen_maq["ultima_fecha"].astype(str)

        for i, row in enumerate(resumen_maq.itertuples(), start=3):
            pct    = f"{round(row.aprobadas / row.total * 100, 1)}%" if row.total > 0 else "0%"
            fill_r = PatternFill("solid", start_color="EBF5FB") if i % 2 == 0 else None
            vals   = [row.maquina, int(row.total), int(row.aprobadas),
                      int(row.con_obs), int(row.rechazadas), pct,
                      str(row.ultimo_insp) if row.ultimo_insp else "", str(row.ultima_fecha)]
            for ci, v in enumerate(vals, start=1):
                c = ws3.cell(i, ci, v)
                c.font = ft_normal; c.border = borde
                c.alignment = izq if ci in (1, 7) else centro
                if fill_r: c.fill = fill_r
            ws3.row_dimensions[i].height = 18
    ws3.freeze_panes = "A3"

    # ── HOJA 4: RANKING NC ───────────────────────────────────────────────────
    ws4 = wb.create_sheet("Ranking NC")
    ws4.merge_cells("A1:E1")
    ws4["A1"] = "Ranking de No Conformidades por Ítem"
    ws4["A1"].font = ft_titulo; ws4["A1"].fill = fill_titulo; ws4["A1"].alignment = centro
    ws4.row_dimensions[1].height = 26

    hdrs4   = ["SECCIÓN", "DESCRIPCIÓN DEL ÍTEM", "DÍA", "# NC", "% NC"]
    anchos4 = [30, 65, 8, 8, 10]
    for ci, (h, w) in enumerate(zip(hdrs4, anchos4), start=1):
        c = ws4.cell(2, ci, h)
        c.font = ft_header; c.fill = PatternFill("solid", start_color="922B21")
        c.alignment = centro; c.border = borde
        ws4.column_dimensions[get_column_letter(ci)].width = w
    ws4.row_dimensions[2].height = 24

    if not df_all_items.empty:
        total_insp = len(df_inspecciones)
        ranking_nc = (
            df_all_items.groupby(["seccion", "descripcion", "dia"], as_index=False)
            .agg(num_nc=("resultado", lambda x: (x == "NC").sum()))
            .sort_values("num_nc", ascending=False)
        )
        for i, row in enumerate(ranking_nc.itertuples(), start=3):
            pct    = f"{round(row.num_nc / total_insp * 100, 1)}%" if total_insp > 0 else "0%"
            fill_r4 = PatternFill("solid", start_color="FADBD8") if row.num_nc > 0 and i % 2 == 0 else (
                      PatternFill("solid", start_color="EBF5FB") if i % 2 == 0 else None)
            vals = [row.seccion, row.descripcion, str(row.dia), int(row.num_nc), pct]
            for ci, v in enumerate(vals, start=1):
                c = ws4.cell(i, ci, v)
                c.font = ft_nc if row.num_nc > 0 and ci == 4 else ft_normal
                c.border = borde
                c.alignment = izq if ci == 2 else centro
                if fill_r4: c.fill = fill_r4
            ws4.row_dimensions[i].height = 18
    ws4.freeze_panes = "A3"

    # ── HOJA 5: POR INSPECTOR ────────────────────────────────────────────────
    ws5 = wb.create_sheet("Por Inspector")
    ws5.merge_cells("A1:F1")
    ws5["A1"] = "Resumen por Inspector / Trabajador"
    ws5["A1"].font = ft_titulo; ws5["A1"].fill = fill_titulo; ws5["A1"].alignment = centro
    ws5.row_dimensions[1].height = 26

    hdrs5   = ["TRABAJADOR","TOTAL INSP.","✅ APROBADAS","⚠️ CON OBS.","❌ RECHAZADAS","% APROBACIÓN"]
    anchos5 = [30, 12, 14, 14, 14, 14]
    for ci, (h, w) in enumerate(zip(hdrs5, anchos5), start=1):
        c = ws5.cell(2, ci, h)
        c.font = ft_header; c.fill = PatternFill("solid", start_color="1A5276")
        c.alignment = centro; c.border = borde
        ws5.column_dimensions[get_column_letter(ci)].width = w
    ws5.row_dimensions[2].height = 24

    if not df_inspecciones.empty and "trabajador" in df_inspecciones.columns:
        df_trab_filtrado = df_inspecciones[
            df_inspecciones["trabajador"].notna() & (df_inspecciones["trabajador"].str.strip() != "")
        ]
        if not df_trab_filtrado.empty:
            resumen_insp = df_trab_filtrado.groupby("trabajador", as_index=False).agg(
                total      =("trabajador", "count"),
                aprobadas  =("estado",     lambda x: x.str.contains("Aprobada",      na=False).sum()),
                con_obs    =("estado",     lambda x: x.str.contains("Observaciones", na=False).sum()),
                rechazadas =("estado",     lambda x: x.str.contains("Rechazada",     na=False).sum()),
            ).sort_values("total", ascending=False)

            for i, row in enumerate(resumen_insp.itertuples(), start=3):
                pct    = f"{round(row.aprobadas / row.total * 100, 1)}%" if row.total > 0 else "0%"
                fill_r = PatternFill("solid", start_color="EBF5FB") if i % 2 == 0 else None
                vals   = [row.trabajador, int(row.total), int(row.aprobadas),
                          int(row.con_obs), int(row.rechazadas), pct]
                for ci, v in enumerate(vals, start=1):
                    c = ws5.cell(i, ci, v)
                    c.font = ft_normal; c.border = borde
                    c.alignment = izq if ci == 1 else centro
                    if fill_r: c.fill = fill_r
                ws5.row_dimensions[i].height = 18
    ws5.freeze_panes = "A3"

    # ── HOJA 6: GRÁFICAS ─────────────────────────────────────────────────────
    try:
        from openpyxl.chart import PieChart, Reference
        from openpyxl.chart.series import DataPoint

        ws6 = wb.create_sheet("Gráficas")
        ws6["A1"] = "Estado"; ws6["B1"] = "Cantidad"
        ws6["A1"].font = ft_header; ws6["B1"].font = ft_header
        ws6["A1"].fill = fill_header; ws6["B1"].fill = fill_header

        estados_g = ["Aprobada", "Con Observaciones", "Rechazada"]
        for i, est in enumerate(estados_g, start=2):
            cnt = len(df_inspecciones[df_inspecciones["estado"].str.contains(est, na=False)]) \
                  if "estado" in df_inspecciones.columns else 0
            ws6.cell(i, 1, est).border = borde
            ws6.cell(i, 2, cnt).border  = borde

        pie = PieChart()
        pie.title  = "Distribución por Estado"
        pie.style  = 10
        labels     = Reference(ws6, min_col=1, min_row=2, max_row=4)
        data       = Reference(ws6, min_col=2, min_row=1, max_row=4)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.width  = 15; pie.height = 12
        colores_g  = ["2ECC71", "F39C12", "E74C3C"]
        for idx, color in enumerate(colores_g):
            pt = DataPoint(idx=idx)
            pt.graphicalProperties.solidFill = color
            pie.series[0].dPt.append(pt)
        ws6.add_chart(pie, "D1")
        for col_l, w in zip(["A","B"], [22, 10]):
            ws6.column_dimensions[col_l].width = w
    except Exception:
        pass

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ==================== MAIN ====================
def main():
    st.markdown("""
    <div class="main-header">
        <h1>🔧 INSPECCIONES PREOPERACIONALES</h1>
        <p>Registro y seguimiento de inspecciones de equipos — SCA ZF</p>
    </div>
    """, unsafe_allow_html=True)

    if "db" not in st.session_state:
        st.session_state.db = DB()
    if "editando_id" not in st.session_state:
        st.session_state.editando_id = None

    db = st.session_state.db

    tab1, tab2, tab3 = st.tabs(["📝 Nueva Inspección", "🔍 Historial y Reportes", "📊 Dashboard"])

    # =========================================================================
    # TAB 1 – NUEVA INSPECCIÓN (SEMANAL)
    # =========================================================================
    with tab1:
        st.markdown("### Registrar Nueva Inspección Preoperacional Semanal")

        # ── 1. Datos del equipo ───────────────────────────────────────────────
        st.markdown("<div class='seccion-titulo'>🏭 1. DATOS DEL EQUIPO</div>", unsafe_allow_html=True)
        d1, d2, d3, d4, d5, d6, d7 = st.columns(7)
        with d1:
            fecha_ini_insp = st.date_input("📅 Fecha Inicio", datetime.now(), key="n_fecha_ini")
        with d2:
            fecha_fin_insp = st.date_input("📅 Fecha Fin",
                                            datetime.now() + timedelta(days=6),
                                            key="n_fecha_fin")
        with d3:
            maquina_sel = st.selectbox("⚙️ Máquina", MAQUINAS, key="n_maquina")
        with d4:
            modelo_inp  = st.text_input("Modelo",       placeholder="Ej: ZF-200X", key="n_modelo")
        with d5:
            marca_inp   = st.text_input("Marca",        placeholder="Ej: SCA",     key="n_marca")
        with d6:
            placa_inp   = st.text_input("Placa / Serie", placeholder="Ej: MQ-001", key="n_placa")

        # Validar rango de fechas
        if fecha_fin_insp < fecha_ini_insp:
            st.error("⚠️ La fecha de fin no puede ser anterior a la fecha de inicio.")
            return

        # Calcular días entre las fechas (máx 7)
        delta_dias = (fecha_fin_insp - fecha_ini_insp).days + 1
        if delta_dias > 7:
            st.warning("⚠️ El rango seleccionado supera 7 días. Se mostrarán solo los primeros 7.")
            delta_dias = 7

        # Generar lista de días con su nombre corto
        dias_labels = []
        dias_nombres_es = ["Lun", "Mar", "Mier", "Juev", "Vier", "Sáb", "Dom"]
        for i in range(delta_dias):
            d = fecha_ini_insp + timedelta(days=i)
            nombre_dia = dias_nombres_es[d.weekday()]
            dias_labels.append(f"{nombre_dia} {d.strftime('%d/%m')}")

        # ── Selector de días activos ──────────────────────────────────────────
        st.markdown("<div class='seccion-titulo'>📅 2. DÍAS TRABAJADOS EN LA SEMANA</div>", unsafe_allow_html=True)
        st.caption("Marca los días en que la máquina fue operada (se generará una columna por cada día seleccionado)")

        cols_dias = st.columns(len(dias_labels))
        dias_activos = []
        for j, label in enumerate(dias_labels):
            with cols_dias[j]:
                activo = st.checkbox(label, value=True, key=f"n_dia_{j}")
                if activo:
                    dias_activos.append(label)

        if not dias_activos:
            st.warning("Selecciona al menos un día trabajado.")
        else:
            if db.verificar_inspeccion_existente(fecha_ini_insp, maquina_sel):
                st.warning(f"⚠️ Ya existe una inspección para **{maquina_sel}** iniciando el **{fecha_ini_insp}**. Si continúas, se registrará una segunda.")

            # ── Tablas de inspección (una por sección, columnas = días) ──────
            st.markdown("<div class='seccion-titulo'>🔍 3. LISTA DE ACTIVIDADES — ANTES DE SU USO</div>",
                        unsafe_allow_html=True)
            render_tabla_semanal("ANTES DE SU USO", ITEMS_ANTES_USO,
                                  "new", "au", dias_activos)

            render_tabla_semanal("🦺 ELEMENTOS DE PROTECCIÓN PERSONAL", ITEMS_EPP,
                                  "new", "epp", dias_activos)

            render_tabla_semanal("⚡ SEGURIDAD ELÉCTRICA", ITEMS_ELECTRICA,
                                  "new", "elec", dias_activos)

            # ── Datos de control ─────────────────────────────────────────────
            st.markdown("<div class='seccion-titulo'>📋 4. DATOS DE CONTROL — <span class='campo-obligatorio'>* Todos los campos son obligatorios</span></div>",
                        unsafe_allow_html=True)
            c1, c2, c3, c4 = st.columns(4)
            with c1: trabajador_inp = st.text_input("👷 Trabajador *",          placeholder="Nombre del operario",    key="n_trab")
            with c2: revisado_inp   = st.text_input("👤 Revisado por *",        placeholder="Supervisor / Jefe",      key="n_rev")
            with c3: cliente_inp    = st.text_input("🏢 Cliente / Proyecto *",  placeholder="Nombre del proyecto",    key="n_cli")
            with c4: resp_mant_inp  = st.text_input("🔧 Resp. Mantenimiento *", placeholder="Nombre del responsable", key="n_mant")

            e1, e2 = st.columns([1, 3])
            with e1: estado_inp = st.selectbox("🚦 Estado", ESTADOS_INSPECCION, key="n_estado")
            with e2: obs_inp    = st.text_area(
                "💬 Comentarios / Observaciones",
                placeholder="Describa cualquier anomalía. REPORTAR INMEDIATAMENTE al encargado de equipos y al departamento de mantenimiento.",
                height=90, key="n_obs")

            st.divider()
            if st.button("💾 Guardar Inspección Semanal", type="primary",
                          use_container_width=True, key="btn_guardar"):
                errores = validar_datos_control(trabajador_inp, revisado_inp, cliente_inp, resp_mant_inp)
                if not maquina_sel:
                    errores.insert(0, "⚙️ La **Máquina** es obligatoria.")
                if errores:
                    st.error("❌ Por favor completa los siguientes campos obligatorios:")
                    for err in errores:
                        st.markdown(f"- {err}")
                else:
                    items_form = construir_items_semanal("new", dias_activos)
                    datos = {
                        "fecha_inicio": fecha_ini_insp,
                        "fecha_fin":    fecha_fin_insp,
                        "maquina":      maquina_sel,
                        "modelo":       modelo_inp,
                        "marca":        marca_inp,
                        "placa":        placa_inp,
                        "trabajador":   trabajador_inp.strip(),
                        "revisado_por": revisado_inp.strip(),
                        "cliente_proyecto":          cliente_inp.strip(),
                        "responsable_mantenimiento": resp_mant_inp.strip(),
                        "estado":        estado_inp.split(" ", 1)[1] if " " in estado_inp else estado_inp,
                        "observaciones": obs_inp,
                    }
                    nc_count = sum(1 for it in items_form if it["resultado"] == "NC")
                    if db.guardar_inspeccion(datos, items_form):
                        st.success(
                            f"✅ Inspección semanal guardada — {maquina_sel} | "
                            f"{fecha_ini_insp} → {fecha_fin_insp} | {nc_count} NC detectadas"
                        )
                        if nc_count > 0:
                            st.warning(f"⚠️ Se detectaron **{nc_count} No Conformidades**. Reportar al encargado de mantenimiento.")
                        st.balloons()

    # =========================================================================
    # TAB 2 – HISTORIAL
    # =========================================================================
    with tab2:
        st.markdown("### 🔍 Historial de Inspecciones")

        with st.expander("🛠️ Filtros", expanded=True):
            f1, f2, f3, f4, f5 = st.columns(5)
            with f1: fi    = st.date_input("Desde", datetime.now() - timedelta(days=30), key="h_fi")
            with f2: ff    = st.date_input("Hasta", datetime.now(), key="h_ff")
            with f3:
                maq_opts = ["Todas"] + MAQUINAS
                fm = st.selectbox("Máquina", maq_opts, key="h_fm")
            with f4: ftrab = st.text_input("Trabajador", key="h_trab")
            with f5:
                est_opts = ["Todos"] + [e.split(" ", 1)[1] for e in ESTADOS_INSPECCION]
                fe = st.selectbox("Estado", est_opts, key="h_fe")

        df_hist = db.obtener_inspecciones(
            fi, ff,
            fm    if fm    != "Todas" else None,
            fe    if fe    != "Todos" else None,
            ftrab if ftrab else None
        )

        if not df_hist.empty:
            k1, k2, k3, k4 = st.columns(4)
            k1.metric("Total Inspecciones",       len(df_hist))
            k2.metric("✅ Aprobadas",             len(df_hist[df_hist["estado"].str.contains("Aprobada",      na=False)]))
            k3.metric("⚠️ Con Observaciones",     len(df_hist[df_hist["estado"].str.contains("Observaciones", na=False)]))
            k4.metric("❌ Rechazadas",             len(df_hist[df_hist["estado"].str.contains("Rechazada",     na=False)]))

            st.divider()
            col_e1, col_e2 = st.columns([2, 5])
            with col_e1:
                nombre_rep = st.text_input("Nombre del reporte", value="Inspecciones_Preop", key="rep_nombre")
            with col_e2:
                st.markdown("<br>", unsafe_allow_html=True)
                excel_data = generar_excel(df_hist, db, titulo=nombre_rep)
                st.download_button(
                    "⬇️ Descargar Excel", data=excel_data,
                    file_name=f"{nombre_rep}_{datetime.now(pytz.timezone('America/Bogota')).strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )

            st.divider()
            cols_tabla = ["id","fecha_inicio","fecha_fin","maquina","trabajador","revisado_por",
                          "cliente_proyecto","placa","estado","observaciones"]
            cols_ex = [c for c in cols_tabla if c in df_hist.columns]
            st.dataframe(df_hist[cols_ex], use_container_width=True, hide_index=True)

            st.divider()
            st.subheader("✏️ Ver Detalle / Editar")
            df_hist["_label"] = df_hist.apply(
                lambda r: (
                    f"ID {r['id']} | {r.get('fecha_inicio', r.get('fecha',''))} → "
                    f"{r.get('fecha_fin', r.get('fecha',''))} | {r['maquina']} | "
                    f"{r.get('trabajador','')} | {r.get('estado','')}"
                ),
                axis=1
            )
            sel = st.selectbox("Seleccionar inspección:", df_hist["_label"].tolist(), key="h_sel")

            if sel:
                vid  = int(sel.split(" | ")[0].replace("ID ", ""))
                row  = df_hist[df_hist["id"] == vid].iloc[0]
                editando       = st.session_state.editando_id == vid
                df_items_sel   = db.obtener_items_inspeccion(vid)

                if not editando:
                    # ── MODO VISTA ──
                    c1, c2, c3 = st.columns(3)
                    with c1:
                        st.info(f"**Máquina:** {row['maquina']}")
                        fi_disp = row.get('fecha_inicio', row.get('fecha',''))
                        ff_disp = row.get('fecha_fin',    row.get('fecha',''))
                        st.write(f"**Semana:** {fi_disp} → {ff_disp}")
                        st.write(f"**Modelo:** {row.get('modelo','')}")
                        st.write(f"**Marca:** {row.get('marca','')}")
                        st.write(f"**Placa/Serie:** {row.get('placa','')}")
                    with c2:
                        st.write(f"**Trabajador:** {row.get('trabajador','')}")
                        st.write(f"**Revisado por:** {row.get('revisado_por','')}")
                        st.write(f"**Cliente/Proyecto:** {row.get('cliente_proyecto','')}")
                        st.write(f"**Resp. Mantenimiento:** {row.get('responsable_mantenimiento','')}")
                    with c3:
                        estado_raw = str(row.get("estado",""))
                        color = "🟢" if "Aprobada" in estado_raw else ("🔴" if "Rechazada" in estado_raw else "🟡")
                        st.write(f"**Estado:** {color} {estado_raw}")
                        st.write(f"**Observaciones:** {row.get('observaciones','')}")

                    if not df_items_sel.empty:
                        num_nc = len(df_items_sel[df_items_sel["resultado"] == "NC"])
                        num_c  = len(df_items_sel[df_items_sel["resultado"] == "C"])
                        st.write(f"**Resultado ítems:** 🟢 {num_c} Conformes · 🔴 {num_nc} No Conformes")

                        # Mostrar tabla pivote por día
                        for sec in df_items_sel["seccion"].unique():
                            items_sec = df_items_sel[df_items_sel["seccion"] == sec].copy()
                            st.markdown(f"**{sec}**")
                            dias_unicos = items_sec["dia"].unique().tolist()
                            col_r = [4] + [1] * len(dias_unicos)
                            hh = st.columns(col_r)
                            hh[0].markdown("**Ítem**")
                            for j, d in enumerate(dias_unicos):
                                hh[j+1].markdown(f"<div class='dia-header'>{d}</div>", unsafe_allow_html=True)
                            for item_num in items_sec["item_numero"].unique():
                                fila_i = items_sec[items_sec["item_numero"] == item_num]
                                desc_i = fila_i.iloc[0]["descripcion"]
                                rr = st.columns(col_r)
                                with rr[0]:
                                    st.markdown(f"<div class='item-label'>{item_num}. {desc_i}</div>",
                                                unsafe_allow_html=True)
                                for j, d in enumerate(dias_unicos):
                                    res_d = fila_i[fila_i["dia"] == d]["resultado"].values
                                    val_d = res_d[0] if len(res_d) > 0 else "—"
                                    with rr[j+1]:
                                        st.write(badge_resultado(val_d))

                    bc1, bc2 = st.columns(2)
                    with bc1:
                        if st.button("✏️ Editar esta inspección", key=f"eb_{vid}", use_container_width=True):
                            st.session_state.editando_id = vid
                            st.rerun()
                    with bc2:
                        if st.button("🗑️ Eliminar", key=f"del_{vid}", use_container_width=True):
                            if db.eliminar_inspeccion(vid):
                                st.success("✅ Inspección eliminada correctamente.")
                                st.rerun()

                else:
                    # ── MODO EDICIÓN ──
                    st.markdown(f"#### ✏️ Editando inspección ID {vid}")
                    st.caption("Los campos marcados con * son obligatorios")

                    # Reconstruir días activos previos desde los ítems guardados
                    dias_previos = []
                    if not df_items_sel.empty:
                        dias_previos = df_items_sel["dia"].unique().tolist()

                    prev_vals = {}
                    if not df_items_sel.empty:
                        sec_map = {
                            "ANTES DE SU USO":                 ("au",   ITEMS_ANTES_USO),
                            "ELEMENTOS DE PROTECCIÓN PERSONAL": ("epp",  ITEMS_EPP),
                            "SEGURIDAD ELÉCTRICA":             ("elec", ITEMS_ELECTRICA),
                        }
                        for _, it in df_items_sel.iterrows():
                            sec_key_map = sec_map.get(it["seccion"])
                            if sec_key_map:
                                sec_k, sec_list = sec_key_map
                                idx_i = int(it["item_numero"]) - 1
                                dia_v = it.get("dia", "General")
                                prev_vals[f"edit_{vid}_{sec_k}_{idx_i}_{dia_v}"] = it["resultado"]

                    # Datos del equipo
                    st.markdown("<div class='seccion-titulo'>🏭 DATOS DEL EQUIPO</div>", unsafe_allow_html=True)
                    ec1, ec2, ec3, ec4, ec5, ec6, ec7 = st.columns(7)
                    with ec1:
                        fi_prev = row.get("fecha_inicio", date.today())
                        e_fecha_ini = st.date_input("📅 Fecha Inicio", value=fi_prev, key=f"efi_{vid}")
                    with ec2:
                        ff_prev = row.get("fecha_fin", date.today())
                        e_fecha_fin = st.date_input("📅 Fecha Fin",   value=ff_prev, key=f"eff_{vid}")
                    with ec3:
                        maq_idx = MAQUINAS.index(row["maquina"]) if row["maquina"] in MAQUINAS else 0
                        e_maq   = st.selectbox("⚙️ Máquina", MAQUINAS, index=maq_idx, key=f"em_{vid}")
                    with ec4:
                        e_modelo = st.text_input("Modelo", value=str(row.get("modelo","") or ""), key=f"emod_{vid}")
                    with ec5:
                        e_marca  = st.text_input("Marca",  value=str(row.get("marca", "") or ""), key=f"emarca_{vid}")
                    with ec6:
                        e_placa  = st.text_input("Placa",  value=str(row.get("placa", "") or ""), key=f"eplaca_{vid}")

                    # Días activos en edición
                    if e_fecha_fin < e_fecha_ini:
                        st.error("La fecha fin no puede ser anterior a la fecha inicio.")
                    else:
                        delta_e = (e_fecha_fin - e_fecha_ini).days + 1
                        if delta_e > 7: delta_e = 7
                        dias_nombres_es = ["Lun", "Mar", "Mier", "Juev", "Vier", "Sáb", "Dom"]
                        dias_labels_e = []
                        for i in range(delta_e):
                            d = e_fecha_ini + timedelta(days=i)
                            dias_labels_e.append(f"{dias_nombres_es[d.weekday()]} {d.strftime('%d/%m')}")

                        st.markdown("<div class='seccion-titulo'>📅 DÍAS TRABAJADOS</div>", unsafe_allow_html=True)
                        cols_de = st.columns(len(dias_labels_e))
                        dias_activos_e = []
                        for j, label in enumerate(dias_labels_e):
                            with cols_de[j]:
                                activo = st.checkbox(label, value=(label in dias_previos or True),
                                                     key=f"edit_{vid}_dia_{j}")
                                if activo:
                                    dias_activos_e.append(label)

                        if dias_activos_e:
                            st.caption("Selecciona **C** = Cumple · **NC** = No Cumple · **N/A** = No Aplica")
                            render_tabla_semanal("ANTES DE SU USO",                  ITEMS_ANTES_USO,
                                                  f"edit_{vid}", "au",   dias_activos_e, prev_vals)
                            render_tabla_semanal("ELEMENTOS DE PROTECCIÓN PERSONAL", ITEMS_EPP,
                                                  f"edit_{vid}", "epp",  dias_activos_e, prev_vals)
                            render_tabla_semanal("SEGURIDAD ELÉCTRICA",              ITEMS_ELECTRICA,
                                                  f"edit_{vid}", "elec", dias_activos_e, prev_vals)

                    # Datos de control
                    st.markdown("<div class='seccion-titulo'>📋 DATOS DE CONTROL — <span class='campo-obligatorio'>* Todos los campos son obligatorios</span></div>",
                                unsafe_allow_html=True)
                    ee1, ee2, ee3, ee4 = st.columns(4)
                    with ee1:
                        e_trab = st.text_input("👷 Trabajador *",       value=str(row.get("trabajador","")                or ""), key=f"etrab_{vid}")
                    with ee2:
                        e_rev  = st.text_input("👤 Revisado por *",     value=str(row.get("revisado_por","")              or ""), key=f"erev_{vid}")
                    with ee3:
                        e_cli  = st.text_input("🏢 Cliente/Proyecto *", value=str(row.get("cliente_proyecto","")          or ""), key=f"ecli_{vid}")
                    with ee4:
                        e_mant = st.text_input("🔧 Resp. Mant. *",      value=str(row.get("responsable_mantenimiento","") or ""), key=f"emant_{vid}")

                    estados_l  = [e.split(" ", 1)[1] for e in ESTADOS_INSPECCION]
                    est_actual = str(row.get("estado") or "Aprobada")
                    est_idx = 0
                    for idx_e, est_opt in enumerate(estados_l):
                        if est_actual in est_opt or est_opt in est_actual:
                            est_idx = idx_e
                            break

                    ef1, ef2 = st.columns([1, 3])
                    with ef1:
                        e_estado = st.selectbox("🚦 Estado", ESTADOS_INSPECCION, index=est_idx, key=f"eest_{vid}")
                    with ef2:
                        e_obs = st.text_area("💬 Observaciones", value=str(row.get("observaciones","") or ""),
                                              key=f"eobs_{vid}", height=80)

                    st.divider()
                    sg1, sg2 = st.columns(2)
                    with sg1:
                        guardar_edit = st.button("💾 Guardar Cambios", type="primary",
                                                  key=f"guardar_edit_{vid}", use_container_width=True)
                    with sg2:
                        cancelar_edit = st.button("❌ Cancelar Edición",
                                                   key=f"cancelar_edit_{vid}", use_container_width=True)

                    if guardar_edit:
                        errores_edit = validar_datos_control(e_trab, e_rev, e_cli, e_mant)
                        if errores_edit:
                            st.error("❌ Por favor completa los siguientes campos obligatorios:")
                            for err in errores_edit:
                                st.markdown(f"- {err}")
                        else:
                            items_edit = construir_items_semanal(f"edit_{vid}", dias_activos_e)
                            datos_edit = {
                                "fecha_inicio": e_fecha_ini, "fecha_fin": e_fecha_fin,
                                "maquina": e_maq, "modelo": e_modelo,
                                "marca": e_marca, "placa": e_placa,
                                "trabajador": e_trab.strip(), "revisado_por": e_rev.strip(),
                                "cliente_proyecto": e_cli.strip(), "responsable_mantenimiento": e_mant.strip(),
                                "estado": e_estado.split(" ", 1)[1] if " " in e_estado else e_estado,
                                "observaciones": e_obs,
                            }
                            if db.actualizar_inspeccion(vid, datos_edit, items_edit):
                                st.success("✅ Inspección actualizada correctamente.")
                                st.session_state.editando_id = None
                                st.rerun()

                    if cancelar_edit:
                        st.session_state.editando_id = None
                        st.rerun()

        else:
            st.warning("No hay inspecciones con los filtros seleccionados.")

    # =========================================================================
    # TAB 3 – DASHBOARD
    # =========================================================================
    with tab3:
        st.markdown("### 📊 Dashboard de Inspecciones")
        try:
            import plotly.express as px

            col_r1, _ = st.columns([2, 4])
            with col_r1:
                rango = st.date_input(
                    "Período",
                    value=(datetime.now().replace(day=1), datetime.now()),
                    key="dash_rango"
                )

            if not (isinstance(rango, (list, tuple)) and len(rango) == 2):
                st.info("Selecciona un rango de fechas completo.")
                return

            df_s = db.stats_dashboard(rango[0], rango[1])
            if df_s.empty:
                st.info("No hay datos en este período.")
                return

            total  = len(df_s)
            apro   = len(df_s[df_s["estado"].str.contains("Aprobada",      na=False)])
            obs_c  = len(df_s[df_s["estado"].str.contains("Observaciones", na=False)])
            rech_c = len(df_s[df_s["estado"].str.contains("Rechazada",     na=False)])
            pct    = round(apro / total * 100) if total > 0 else 0
            total_nc = int(df_s["num_nc"].sum())

            k1, k2, k3, k4, k5 = st.columns(5)
            k1.metric("🔧 Total Inspecciones",  total)
            k2.metric("✅ Aprobadas",            apro,     f"{pct}%")
            k3.metric("⚠️ Con Observaciones",   obs_c)
            k4.metric("❌ Rechazadas",           rech_c)
            k5.metric("🔴 Total NC detectadas", total_nc)

            st.divider()
            g1, g2 = st.columns(2)
            with g1:
                st.markdown("#### Distribución por Estado")
                est_c_df = df_s["estado"].value_counts().reset_index()
                est_c_df.columns = ["estado","cantidad"]
                colores_est = {"Aprobada":"#2ecc71","Con Observaciones":"#f39c12","Rechazada":"#e74c3c"}
                fig1 = px.pie(est_c_df, values="cantidad", names="estado", hole=0.45,
                              color="estado", color_discrete_map=colores_est)
                fig1.update_layout(margin=dict(t=10,b=10), height=300)
                st.plotly_chart(fig1, use_container_width=True)
            with g2:
                st.markdown("#### Inspecciones por Semana (Fecha Inicio)")
                df_dia = df_s.groupby("fecha").size().reset_index(name="inspecciones")
                fig2 = px.bar(df_dia, x="fecha", y="inspecciones",
                              color_discrete_sequence=["#2c5364"], text="inspecciones")
                fig2.update_traces(textposition="outside")
                fig2.update_layout(margin=dict(t=10,b=10), height=300, xaxis_title="", yaxis_title="Inspecciones")
                st.plotly_chart(fig2, use_container_width=True)

            st.divider()
            g3, g4 = st.columns(2)
            with g3:
                st.markdown("#### Inspecciones por Máquina")
                df_maq = df_s.groupby("maquina").size().reset_index(name="inspecciones").sort_values("inspecciones")
                fig3 = px.bar(df_maq, x="inspecciones", y="maquina", orientation="h",
                              color="inspecciones", color_continuous_scale="Blues", text="inspecciones")
                fig3.update_traces(textposition="outside")
                fig3.update_layout(margin=dict(t=10,b=10), height=max(250,len(df_maq)*45),
                                   coloraxis_showscale=False, yaxis_title="", xaxis_title="Inspecciones")
                st.plotly_chart(fig3, use_container_width=True)
            with g4:
                st.markdown("#### NC Promedio por Máquina")
                df_nc_maq = df_s.groupby("maquina").agg(prom_nc=("num_nc","mean")).reset_index().sort_values("prom_nc")
                df_nc_maq["prom_nc"] = df_nc_maq["prom_nc"].round(1)
                fig4 = px.bar(df_nc_maq, x="prom_nc", y="maquina", orientation="h",
                              color="prom_nc", color_continuous_scale="Reds", text="prom_nc")
                fig4.update_traces(textposition="outside")
                fig4.update_layout(margin=dict(t=10,b=10), height=max(250,len(df_nc_maq)*45),
                                   coloraxis_showscale=False, yaxis_title="", xaxis_title="NC promedio")
                st.plotly_chart(fig4, use_container_width=True)

            st.divider()
            g5, g6 = st.columns(2)
            with g5:
                st.markdown("#### % Aprobación por Máquina")
                resumen_apr = df_s.groupby("maquina", as_index=False).agg(
                    aprobadas=("estado", lambda x: x.str.contains("Aprobada", na=False).sum()),
                    total    =("estado", "count"),
                )
                resumen_apr["pct_apro"] = (resumen_apr["aprobadas"] / resumen_apr["total"] * 100).round(1)
                resumen_apr = resumen_apr.sort_values("pct_apro")
                fig5 = px.bar(resumen_apr, x="pct_apro", y="maquina", orientation="h",
                              color="pct_apro", color_continuous_scale="Greens",
                              text="pct_apro", range_x=[0,100])
                fig5.update_traces(texttemplate="%{text}%", textposition="outside")
                fig5.update_layout(margin=dict(t=10,b=10), height=max(250,len(resumen_apr)*45),
                                   coloraxis_showscale=False, yaxis_title="", xaxis_title="% Aprobación")
                st.plotly_chart(fig5, use_container_width=True)
            with g6:
                st.markdown("#### 📅 Inspecciones por Día de la Semana")
                df_s["dia_semana"] = pd.to_datetime(df_s["fecha"]).dt.day_name()
                orden = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
                nombres_es = {"Monday":"Lunes","Tuesday":"Martes","Wednesday":"Miércoles",
                              "Thursday":"Jueves","Friday":"Viernes","Saturday":"Sábado","Sunday":"Domingo"}
                df_sem = df_s.groupby("dia_semana").size().reset_index(name="inspecciones")
                df_sem["orden"] = df_sem["dia_semana"].map({d: i for i, d in enumerate(orden)})
                df_sem = df_sem.sort_values("orden")
                df_sem["dia_es"] = df_sem["dia_semana"].map(nombres_es)
                fig6 = px.bar(df_sem, x="dia_es", y="inspecciones",
                              color="inspecciones", color_continuous_scale="Oranges", text="inspecciones")
                fig6.update_traces(textposition="outside")
                fig6.update_layout(margin=dict(t=10,b=10), height=300,
                                   coloraxis_showscale=False, xaxis_title="", yaxis_title="Inspecciones")
                st.plotly_chart(fig6, use_container_width=True)

            st.divider()
            st.markdown("#### 🏆 Ranking de Inspectores")
            df_insp = df_s[
                df_s["trabajador"].notna() & (df_s["trabajador"].str.strip() != "")
            ].groupby("trabajador", as_index=False).agg(
                inspecciones=("trabajador","count"),
                aprobadas   =("estado", lambda x: x.str.contains("Aprobada",      na=False).sum()),
                con_obs     =("estado", lambda x: x.str.contains("Observaciones", na=False).sum()),
                rechazadas  =("estado", lambda x: x.str.contains("Rechazada",     na=False).sum()),
                total_nc    =("num_nc","sum"),
            ).sort_values("inspecciones", ascending=False)
            df_insp["% Aprobación"] = (df_insp["aprobadas"] / df_insp["inspecciones"] * 100).round(1).astype(str) + "%"
            df_insp["total_nc"]     = df_insp["total_nc"].astype(int)
            df_insp.columns = ["Inspector","Total","✅ Aprob.","⚠️ Obs.","❌ Rech.","🔴 NC Total","% Aprobación"]
            st.dataframe(df_insp, use_container_width=True, hide_index=True)

        except ImportError:
            st.warning("Instala plotly: `pip install plotly`")
        except Exception as e:
            st.error(f"Error en dashboard: {e}")


if __name__ == "__main__":
    main()
